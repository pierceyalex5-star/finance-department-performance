#!/usr/bin/env python3
import io, json, re, sys, urllib.request
from datetime import datetime
from openpyxl import load_workbook

PATH = sys.argv[1] if len(sys.argv) > 1 else 'trend-data.js'
XLSX_URL = 'https://energir.com/files/energir_common/Evolution_prix_en.xlsx'
SOURCE_URL = 'https://energir.com/en/business/customer-centre/billing-and-pricing/pricing'

# Official monthly values visible in Énergir's published pricing table.
# These override the workbook for the period used by the landed-cost selector.
# tuple: date, reference price ¢/m3, 1-year market $/GJ, monthly market $/GJ
OFFICIAL_RECENT = [
    ('2025-01-01',10.003,4.39,4.73),('2025-02-01',14.663,4.45,4.64),
    ('2025-03-01',19.476,5.33,5.23),('2025-04-01',22.582,5.43,5.15),
    ('2025-05-01',19.362,4.46,3.66),('2025-06-01',18.528,4.46,3.57),
    ('2025-07-01',17.884,4.61,3.69),('2025-08-01',17.808,4.35,3.56),
    ('2025-09-01',16.406,4.24,3.24),('2025-10-01',16.785,4.53,3.65),
    ('2025-11-01',17.846,4.78,4.11),('2025-12-01',18.907,5.06,5.74),
    ('2026-01-01',17.770,4.62,5.50),('2026-02-01',16.937,5.61,10.70),
    ('2026-03-01',18.339,4.15,3.78),('2026-04-01',15.308,4.34,3.61),
    ('2026-05-01',15.611,3.97,3.06),('2026-06-01',15.156,4.00,3.33),
    ('2026-07-01',15.346,4.18,3.69),
]


def parse_period(value, year_hint):
    if isinstance(value, datetime):
        # Some workbook tabs preserve an old Excel year in the cell. Use the sheet
        # year when it clearly disagrees with the tab being parsed.
        year = year_hint if value.year != year_hint else value.year
        return f'{year:04d}-{value.month:02d}-01'
    s = str(value or '').strip()
    if not s:
        return None
    for fmt in ('%B %Y', '%m-%Y', '%m/%Y', '%Y-%m'):
        try:
            dt = datetime.strptime(s, fmt)
            return f'{year_hint if dt.year != year_hint else dt.year:04d}-{dt.month:02d}-01'
        except ValueError:
            pass
    m = re.search(r'(\d{1,2})[-/]?(\d{4})', s)
    if m:
        year = int(m.group(2))
        return f'{year_hint if year != year_hint else year:04d}-{int(m.group(1)):02d}-01'
    # Month-name-only cells are also valid inside a year worksheet.
    for fmt in ('%B','%b'):
        try:
            dt=datetime.strptime(s,fmt)
            return f'{year_hint:04d}-{dt.month:02d}-01'
        except ValueError:
            pass
    return None


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('*', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def load_existing(path):
    text = open(path, encoding='utf-8').read().strip()
    prefix = 'window.FI_TREND_DATA='
    if not text.startswith(prefix):
        raise ValueError('Unexpected trend-data.js format')
    body = text[len(prefix):]
    if body.endswith(';'):
        body = body[:-1]
    return json.loads(body)


def fetch_energir_workbook():
    req = urllib.request.Request(XLSX_URL, headers={'User-Agent': 'fastener-intelligence-github-pages/5.1'})
    with urllib.request.urlopen(req, timeout=45) as r:
        blob = r.read()
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    regulated, market_1y, market_monthly = [], [], []
    for ws in wb.worksheets:
        try:
            year_hint = int(ws.title)
        except Exception:
            continue
        if year_hint < 2021:
            continue
        for row in ws.iter_rows(min_row=4, values_only=True):
            date = parse_period(row[0] if len(row) > 0 else None, year_hint)
            if not date:
                continue
            reg_cents = num(row[2] if len(row) > 2 else None)
            one_year = num(row[3] if len(row) > 3 else None)
            monthly = num(row[4] if len(row) > 4 else None)
            if reg_cents is not None:
                regulated.append([date, reg_cents])
            if one_year is not None:
                market_1y.append([date, one_year])
            if monthly is not None:
                market_monthly.append([date, monthly])
    return regulated, market_1y, market_monthly


def overlay_recent(regulated, one_year, monthly):
    r={d:v for d,v in regulated}; y={d:v for d,v in one_year}; m={d:v for d,v in monthly}
    for d,rv,yv,mv in OFFICIAL_RECENT:
        r[d]=rv; y[d]=yv; m[d]=mv
    return [[d,r[d]] for d in sorted(r)],[[d,y[d]] for d in sorted(y)],[[d,m[d]] for d in sorted(m)]


def meta(name, units, relevance, data):
    return {
        'name': name,
        'group': 'Energy',
        'units': units,
        'frequency': 'Monthly',
        'source': 'Énergir natural gas price history / published monthly pricing table',
        'relevance': relevance,
        'url': SOURCE_URL,
        'data': sorted(data),
        'error': None if data else 'Énergir source returned no numeric observations.'
    }

obj = load_existing(PATH)
try:
    regulated, one_year, monthly = fetch_energir_workbook()
except Exception as e:
    print(f'WARN Energir workbook refresh: {e}', file=sys.stderr)
    regulated, one_year, monthly = [], [], []
regulated, one_year, monthly = overlay_recent(regulated, one_year, monthly)
obj.setdefault('series', {})['QC_NATGAS'] = {
    **meta('Québec Natural Gas Supply Price – Énergir', 'cents per cubic metre', 'Québec plant energy / heat-treatment supply cost', regulated),
    'id': 'QC_NATGAS'
}
obj['series']['QC_NATGAS_1Y'] = {
    **meta('Québec Natural Gas Market – 1 Year Contract', 'CAD per GJ', 'Forward natural-gas cost signal for Québec industrial users', one_year),
    'id': 'QC_NATGAS_1Y'
}
obj['series']['QC_NATGAS_MONTHLY'] = {
    **meta('Québec Natural Gas Market – Monthly Price', 'CAD per GJ', 'Spot/monthly natural-gas cost signal for Québec industrial users', monthly),
    'id': 'QC_NATGAS_MONTHLY'
}
obj['provider'] = 'FRED / underlying public agencies + Énergir'

with open(PATH, 'w', encoding='utf-8') as f:
    f.write('window.FI_TREND_DATA=')
    json.dump(obj, f, separators=(',', ':'))
    f.write(';\n')

print(f'Energir regulated observations: {len(regulated)}')
print(f'Energir 1Y observations: {len(one_year)}')
print(f'Energir monthly observations: {len(monthly)}')
if regulated:
    print(f'Latest Energir regulated: {regulated[-1][0]} = {regulated[-1][1]} c/m3')
